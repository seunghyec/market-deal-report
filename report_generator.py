#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
트립비토즈 공구 매출 리포트 자동 생성기
사용법: python gonggu_report_generator.py
"""

import sys
import re
import os
import pandas as pd
from datetime import datetime, timedelta

# ────────────────────────────────────────────────
# 1. 데이터 로드 & 전처리
# ────────────────────────────────────────────────

def load_data(filepath):
    xl = pd.ExcelFile(filepath)
    sheet = xl.sheet_names[0]
    df = pd.read_excel(filepath, sheet_name=sheet)

    required = ['예약 상태', '예약일', '체크인', '체크아웃', '박수', '판매가', '객실명', '요금제명(코드)']
    for col in required:
        if col not in df.columns:
            raise ValueError(f"필수 컬럼 없음: {col}")

    df['예약일'] = pd.to_datetime(df['예약일'])
    df['체크인'] = pd.to_datetime(df['체크인'])
    df['체크아웃'] = pd.to_datetime(df['체크아웃'])
    df['박수'] = pd.to_numeric(df['박수'], errors='coerce').fillna(1).astype(int)
    df['판매가'] = pd.to_numeric(df['판매가'], errors='coerce').fillna(0).astype(int)

    # 상태 정규화 (완료 → 확정)
    df['상태'] = df['예약 상태'].apply(
        lambda x: '확정' if str(x).strip() in ('확정', '완료') else '취소'
    )

    # 대기예약 여부
    df['대기예약여부'] = df['객실명'].str.contains('대기예약', na=False)

    # 객실명 정리
    df['객실명_clean'] = df['객실명'].apply(clean_room)

    # 패키지명 정리 (인원수 통합)
    df['패키지명_clean'] = df['요금제명(코드)'].apply(clean_package)

    # 시설 구분 (객실명에서 [] 안 내용 or 키워드)
    df['시설'] = df['객실명'].apply(extract_facility)

    return df

def clean_room(s):
    s = str(s)
    # 이모지 제거
    s = re.sub(r'[^\x00-\x7F\uAC00-\uD7A3\u3131-\u318E\u1100-\u11FF ()[\]·×,./~\-&+_0-9A-Za-z]', '', s)
    # 마켓딜, 대기예약 제거
    s = re.sub(r'(마켓딜|대기예약)', '', s)
    # [클린]/[콘도]/[EAST]/[WEST] 등 대괄호 유지
    s = re.sub(r'\s*-\s*베드타입 랜덤배정', '', s)
    return re.sub(r'\s+', ' ', s).strip()

def clean_package(s):
    """패키지명에서 호텔 무관 카테고리를 자동 감지"""
    s = str(s)
    is_연박 = '연박' in s
    # 인원 패턴: N+N (소노벨 등) 또는 N인 (롯데 등)
    import re as _re
    m = _re.search(r'(\d+\+\d+)', s)
    if m:
        인원 = m.group(1)
    else:
        m = _re.search(r'(\d+)인', s)
        인원 = f'{m.group(1)}인' if m else ''
    # 연박+오션플레이 조합은 별도 처리 (조식 포함 패키지)
    if '오션플레이' in s and is_연박:
        return f'[연박전용] 오션플레이+조식 {인원}'.strip()
    if '아쿠아' in s:
        pkg = '무제한아쿠아'
    elif '오션플레이' in s:
        pkg = '오션플레이'
    elif '올인클루시브' in s:
        pkg = '올인클루시브'
    elif '조식' in s:
        pkg = '조식'
    else:
        s = _re.sub(r'[^\uAC00-\uD7A3\u3131-\u318E\u1100-\u11FF ()\[\]\d]', '', s)
        return _re.sub(r'\s+', ' ', s).strip()
    prefix = '[연박전용] ' if is_연박 else ''
    return f'{prefix}{pkg} {인원}'.strip()

def extract_facility(s):
    s = str(s)
    # [리조트], [호텔] 등 패턴
    m = re.search(r'\[([^\]]+)\]', s)
    if m:
        return m.group(1)
    # 타워 키워드
    if '웨스트타워' in s: return '웨스트타워'
    if '이스트타워' in s: return '이스트타워'
    if '씨원' in s: return '씨원리조트'
    if '라마다' in s: return '라마다프라자'
    return '전체'

# ────────────────────────────────────────────────
# 2. 집계 함수
# ────────────────────────────────────────────────

def agg(df):
    c = df[df['상태'] == '확정']
    x = df[df['상태'] == '취소']
    return {
        '전체건': len(df), '전체박': int(df['박수'].sum()), '전체매출': int(df['판매가'].sum()),
        '확정건': len(c), '확정박': int(c['박수'].sum()), '확정매출': int(c['판매가'].sum()),
        '취소건': len(x), '취소박': int(x['박수'].sum()), '취소매출': int(x['판매가'].sum()),
    }

def cancel_rate(a):
    total = a['확정건'] + a['취소건']
    return round(a['취소건'] / total * 100, 1) if total > 0 else 0

def adr(매출, 박수):
    return round(매출 / 박수) if 박수 > 0 else 0

def fmt(n):
    return f"{int(n):,}"

# ────────────────────────────────────────────────
# 3. HTML 생성 헬퍼
# ────────────────────────────────────────────────

COLORS = {
    'dark': '#1e4d7b',
    'mid': '#2e6da4',
    'light_bg': '#eef3f8',
    'header_bg': '#f0f4f8',
    'col_bg': '#dce8f4',
    'border': '#d0d7e0',
    'red': '#c0392b',
}

def th(text, rowspan=1, colspan=1, extra=''):
    rs = f' rowspan="{rowspan}"' if rowspan > 1 else ''
    cs = f' colspan="{colspan}"' if colspan > 1 else ''
    return (f'<th{rs}{cs} style="background:{COLORS["header_bg"]}; color:{COLORS["dark"]}; '
            f'padding:6px 5px; text-align:center; border:1px solid {COLORS["border"]}; '
            f'white-space:nowrap; {extra}">{text}</th>')

def th_col(text, colspan=1):
    cs = f' colspan="{colspan}"' if colspan > 1 else ''
    return (f'<th{cs} style="background:{COLORS["col_bg"]}; color:{COLORS["dark"]}; '
            f'padding:6px 5px; text-align:center; border:1px solid {COLORS["border"]}; '
            f'white-space:nowrap;">{text}</th>')

def td_r(val, bold=False, red=False):
    color = f'color:{COLORS["red"]};' if red else ''
    fw = 'font-weight:bold;' if bold else ''
    return (f'<td style="padding:6px 5px; border:1px solid {COLORS["border"]}; '
            f'text-align:right; white-space:nowrap; {color}{fw}">{val}</td>')

def td_c(val, bold=False, red=False):
    color = f'color:{COLORS["red"]};' if red else ''
    fw = 'font-weight:bold;' if bold else ''
    return (f'<td style="padding:6px 5px; border:1px solid {COLORS["border"]}; '
            f'text-align:center; white-space:nowrap; {color}{fw}">{val}</td>')

def td_l(val, small=False):
    fs = 'font-size:11px;' if small else ''
    return (f'<td style="padding:6px 5px; border:1px solid {COLORS["border"]}; '
            f'white-space:nowrap; {fs}">{val}</td>')

def tr_total(*tds):
    return (f'<tr style="background:{COLORS["light_bg"]}; font-weight:bold;">'
            + ''.join(tds) + '</tr>')

def section_header(text):
    return (f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="width:100%; margin-top:16px;">'
            f'<tr><td style="background:{COLORS["light_bg"]}; padding:7px 12px; '
            f'font-weight:bold; font-size:12px; color:{COLORS["dark"]}; '
            f'border-left:3px solid {COLORS["mid"]};">▶ {text}</td></tr></table>')

def facility_header(text, color=None):
    bg = color or COLORS['dark']
    return (f'<tr><td colspan="10" style="background:{bg}; color:#fff; '
            f'padding:5px 10px; font-size:12px; font-weight:bold; '
            f'border:1px solid {bg};">{text}</td></tr>')

def table_wrap(inner, margin_top=4):
    return (f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="width:100%; border-collapse:collapse; font-size:12px; '
            f'margin-top:{margin_top}px;">{inner}</table>')

# ────────────────────────────────────────────────
# 4. 섹션별 HTML 빌더
# ────────────────────────────────────────────────

def build_kpi(a, cr):
    def kpi_cell(label, val, is_rate=False, last=False):
        border = '' if last else f'border-right:1px solid {COLORS["border"]};'
        color = COLORS["red"] if is_rate else COLORS["dark"]
        font = '14px' if len(str(val)) > 10 else '20px'
        return (f'<td style="padding:12px 8px; text-align:center; {border}">'
                f'<div style="font-size:11px; color:#666; margin-bottom:4px;">{label}</div>'
                f'<div style="font-size:{font}; font-weight:bold; color:{color};">{val}'
                f'<span style="font-size:12px; color:#666; font-weight:normal;"> '
                f'{"%" if is_rate else ("원" if "," in str(val) else "건" if "건" in label else "박")}'
                f'</span></div></td>')

    inner = ('<tr>'
             + kpi_cell('확정 건 수', fmt(a['확정건']))
             + kpi_cell('확정 박 수', fmt(a['확정박']), last=False)
             + kpi_cell('확정 매출액', fmt(a['확정매출']), last=False)
             + kpi_cell('취소 건 수', fmt(a['취소건']), last=False)
             + kpi_cell('취소율', cr, is_rate=True, last=True)
             + '</tr>')
    return (f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
            f'style="width:100%; border-collapse:collapse; '
            f'border:1px solid {COLORS["border"]}; border-top:none;">{inner}</table>')

def build_overview(a_total, facility_dict):
    """전체 예약 현황 (시설별)"""
    heads = ('<thead><tr>'
             + th('구분', rowspan=2, extra='width:80px;')
             + th_col('전체 예약', 3) + th_col('확정 예약', 3) + th_col('취소 예약', 3)
             + '</tr><tr>'
             + th('건 수') + th('박 수') + th('매출액 (원)')
             + th('건 수') + th('박 수') + th('매출액 (원)')
             + th('건 수') + th('박 수') + th('매출액 (원)')
             + '</tr></thead>')

    def row(name, a, total_row=False):
        cells = (td_l(name)
                 + td_r(fmt(a['전체건'])) + td_r(fmt(a['전체박'])) + td_r(fmt(a['전체매출']))
                 + td_r(fmt(a['확정건'])) + td_r(fmt(a['확정박'])) + td_r(fmt(a['확정매출']))
                 + td_r(fmt(a['취소건'])) + td_r(fmt(a['취소박'])) + td_r(fmt(a['취소매출'])))
        if total_row:
            return f'<tr style="background:{COLORS["light_bg"]}; font-weight:bold;">{cells}</tr>'
        return f'<tr>{cells}</tr>'

    rows = row('전체', a_total, total_row=True)
    for name, fa in facility_dict.items():
        rows += row(name, fa)

    return section_header(f'전체 예약 현황') + table_wrap('<thead>' + heads + '</thead><tbody>' + rows + '</tbody>')

def build_package(df_conf, df_all):
    """패키지별 현황"""
    pkgs = df_conf.groupby('패키지명_clean')
    all_pkgs = df_all.groupby('패키지명_clean')

    heads = ('<thead><tr>'
             + th('패키지', extra='width:110px;')
             + th('건 수') + th('박 수') + th('매출액 (원)')
             + th('ADR (원/박)') + th('건당 평균단가')
             + th('취소건') + th('취소율') + th('매출비중')
             + '</tr></thead>')

    total_conf_sales = df_conf['판매가'].sum()
    rows = ''
    for pkg_name, g in pkgs:
        # 취소건
        try:
            gall = all_pkgs.get_group(pkg_name)
            x_cnt = len(gall[gall['상태'] == '취소'])
            total_cnt = len(gall)
            cr = f"{round(x_cnt/total_cnt*100,1)}%" if total_cnt > 0 else '-'
        except KeyError:
            x_cnt, cr = 0, '-'

        sales = int(g['판매가'].sum())
        nights = int(g['박수'].sum())
        cnt = len(g)
        pct = f"{round(sales/total_conf_sales*100,1)}%" if total_conf_sales > 0 else '-'

        cells = (td_l(pkg_name)
                 + td_r(fmt(cnt)) + td_r(fmt(nights)) + td_r(fmt(sales))
                 + td_r(fmt(adr(sales, nights))) + td_r(fmt(round(sales/cnt) if cnt else 0))
                 + td_r(fmt(x_cnt)) + td_c(cr, red=True) + td_c(pct))
        rows += f'<tr>{cells}</tr>'

    # 합계
    ts = int(df_conf['판매가'].sum())
    tn = int(df_conf['박수'].sum())
    tc = len(df_conf)
    tx = len(df_all[df_all['상태'] == '취소'])
    ttotal = len(df_all)
    tcr = f"{round(tx/ttotal*100,1)}%" if ttotal > 0 else '-'
    total_cells = (td_l('합계')
                   + td_r(fmt(tc)) + td_r(fmt(tn)) + td_r(fmt(ts))
                   + td_r(fmt(adr(ts, tn))) + td_r(fmt(round(ts/tc) if tc else 0))
                   + td_r(fmt(tx)) + td_c(tcr, red=True) + td_c('100.0%'))
    rows += f'<tr style="background:{COLORS["light_bg"]}; font-weight:bold;">{total_cells}</tr>'

    return section_header('패키지별 현황 (확정 기준)') + table_wrap('<thead>' + heads + '</thead><tbody>' + rows + '</tbody>')

def build_monthly(df):
    """체크인 월별 현황"""
    df = df.copy()
    df['월'] = df['체크인'].dt.to_period('M')
    months = sorted(df['월'].unique())

    conf = df[df['상태'] == '확정']
    canc = df[df['상태'] == '취소']

    # 헤더
    month_heads = ''.join(th_col(str(m).replace('-', '년 ') + '월', 3) for m in months)
    sub_heads = ''.join(th('건') + th('박') + th('매출 (원)') for _ in months)

    heads = ('<thead><tr>'
             + th('구분', rowspan=2, extra='width:30px;')
             + month_heads + th_col('합계', 3)
             + '</tr><tr>' + sub_heads + th('건') + th('박') + th('매출 (원)')
             + '</tr></thead>')

    def make_row(name, dfsub, total_row=False):
        cells = td_c(name)
        for m in months:
            g = dfsub[dfsub['월'] == m]
            cells += (td_r(fmt(len(g))) + td_r(fmt(int(g['박수'].sum())))
                      + td_r(fmt(int(g['판매가'].sum()))))
        cells += (td_r(fmt(len(dfsub))) + td_r(fmt(int(dfsub['박수'].sum())))
                  + td_r(fmt(int(dfsub['판매가'].sum()))))
        if total_row:
            return f'<tr style="font-size:11px;">{cells}</tr>'
        return f'<tr style="font-size:11px;">{cells}</tr>'

    body = ('<tbody>'
            + make_row('전체', df)
            + make_row('확정', conf)
            + make_row('취소', canc)
            + '</tbody>')

    return section_header('체크인 월별 현황 (확정 기준)') + table_wrap('<thead>' + heads + '</thead>' + body)

def build_daily(df):
    """일별 판매 현황"""
    df = df.copy()
    df['예약일자'] = df['예약일'].dt.date
    daily = df.groupby('예약일자').agg(
        건수=('판매가', 'count'), 박수=('박수', 'sum'), 매출=('판매가', 'sum')
    ).reset_index().sort_values('예약일자')

    DOW_KR = {'Monday':'월','Tuesday':'화','Wednesday':'수','Thursday':'목',
               'Friday':'금','Saturday':'토','Sunday':'일'}

    heads = ('<thead><tr>'
             + th('날짜') + th('요일') + th('건 수') + th('박 수')
             + th('매출액 (원)') + th('누적 건 수') + th('누적 매출액 (원)')
             + '</tr></thead>')

    rows = ''
    cum_c = cum_a = 0
    for _, r in daily.iterrows():
        cum_c += r['건수']; cum_a += r['매출']
        dow = DOW_KR.get(pd.Timestamp(r['예약일자']).day_name(), '')
        cells = (td_c(r['예약일자'].strftime('%m-%d'))
                 + td_c(dow)
                 + td_r(fmt(r['건수'])) + td_r(fmt(int(r['박수']))) + td_r(fmt(int(r['매출'])))
                 + td_r(fmt(cum_c)) + td_r(fmt(cum_a)))
        rows += f'<tr>{cells}</tr>'

    total_cells = (td_c('합계', bold=True) + td_c('')
                   + td_r(fmt(int(daily['건수'].sum())), bold=True)
                   + td_r(fmt(int(daily['박수'].sum())), bold=True)
                   + td_r(fmt(int(daily['매출'].sum())), bold=True)
                   + td_c('—') + td_c('—'))
    rows += f'<tr style="background:{COLORS["light_bg"]}; font-weight:bold;">{total_cells}</tr>'

    return section_header('일별 판매 현황') + table_wrap('<thead>' + heads + '</thead><tbody>' + rows + '</tbody>')

def build_weekday(df_conf):
    """요일별 현황"""
    DOW_KR = {'Monday':'월','Tuesday':'화','Wednesday':'수','Thursday':'목',
               'Friday':'금','Saturday':'토','Sunday':'일'}
    DOW_ORDER = ['월','화','수','목','금','토','일']

    df_conf = df_conf.copy()
    df_conf['요일_kr'] = df_conf['체크인'].dt.day_name().map(DOW_KR)
    dw = df_conf.groupby('요일_kr').agg(건수=('판매가','count'), 박수=('박수','sum'), 매출=('판매가','sum')).reindex(DOW_ORDER)

    total_sales = int(df_conf['판매가'].sum())
    total_nights = int(df_conf['박수'].sum())

    heads = ('<thead><tr>'
             + th('요일') + th('건 수') + th('박 수') + th('매출액 (원)')
             + th('ADR (원/박)') + th('비중 (매출)')
             + '</tr></thead>')

    rows = ''
    for dow, r in dw.iterrows():
        if pd.isna(r['건수']): continue
        pct = f"{round(r['매출']/total_sales*100,1)}%"
        is_sat = dow == '토'
        cells = (td_c(f'<b>{dow}</b>' if is_sat else dow)
                 + td_r(fmt(int(r['건수']))) + td_r(fmt(int(r['박수'])))
                 + td_r(fmt(int(r['매출'])))
                 + td_r(fmt(adr(int(r['매출']), int(r['박수']))))
                 + td_c(f'<b>{pct}</b>' if is_sat else pct))
        rows += f'<tr>{cells}</tr>'

    total_cells = (td_c('합계')
                   + td_r(fmt(len(df_conf))) + td_r(fmt(total_nights))
                   + td_r(fmt(total_sales))
                   + td_r(fmt(adr(total_sales, total_nights)))
                   + td_c('100.0%'))
    rows += f'<tr style="background:{COLORS["light_bg"]}; font-weight:bold;">{total_cells}</tr>'

    return section_header('체크인 요일별 현황 (확정 기준)') + table_wrap('<thead>' + heads + '</thead><tbody>' + rows + '</tbody>')

def build_room_analysis(df_all, df_conf, df_canc, facility_order=None):
    """시설별 객실 분석"""
    facilities = facility_order or sorted(df_all['시설'].unique())

    heads = ('<thead><tr>'
             + th('객실명', rowspan=2, extra='width:170px;')
             + th_col('전체 예약', 3) + th_col('확정 예약', 3) + th_col('취소 예약', 3)
             + '</tr><tr>'
             + th('건') + th('박') + th('매출액 (원)')
             + th('건') + th('박') + th('매출액 (원)')
             + th('건') + th('박') + th('매출액 (원)')
             + '</tr></thead>')

    fac_colors = [COLORS['dark'], COLORS['mid'], '#3a7ebf', '#4a8ecf']
    rows = ''
    for fi, fac in enumerate(facilities):
        fa_all = df_all[df_all['시설'] == fac]
        fa_conf = df_conf[df_conf['시설'] == fac]
        fa_canc = df_canc[df_canc['시설'] == fac]

        rows += facility_header(f'▸ {fac}', fac_colors[fi % len(fac_colors)])

        rooms = sorted(fa_all['객실명_clean'].unique())
        for room in rooms:
            ra = fa_all[fa_all['객실명_clean'] == room]
            rc = fa_conf[fa_conf['객실명_clean'] == room]
            rx = fa_canc[fa_canc['객실명_clean'] == room]

            def s(d): return int(d['박수'].sum()), int(d['판매가'].sum())
            an, am = s(ra); cn, cm = s(rc); xn, xm = s(rx)

            cells = (td_l(room, small=True)
                     + td_r(fmt(len(ra))) + td_r(fmt(an)) + td_r(fmt(am))
                     + td_r(fmt(len(rc))) + td_r(fmt(cn)) + td_r(fmt(cm))
                     + td_r(fmt(len(rx))) + td_r(fmt(xn))
                     + td_r(fmt(xm) if xm > 0 else '—'))
            rows += f'<tr>{cells}</tr>'

        # 소계
        def ss(d): return len(d), int(d['박수'].sum()), int(d['판매가'].sum())
        atc, atn, atm = ss(fa_all); ctc, ctn, ctm = ss(fa_conf); xtc, xtn, xtm = ss(fa_canc)
        sub_cells = (td_l(f'{fac} 소계')
                     + td_r(fmt(atc), bold=True) + td_r(fmt(atn), bold=True) + td_r(fmt(atm), bold=True)
                     + td_r(fmt(ctc), bold=True) + td_r(fmt(ctn), bold=True) + td_r(fmt(ctm), bold=True)
                     + td_r(fmt(xtc), bold=True) + td_r(fmt(xtn), bold=True) + td_r(fmt(xtm), bold=True))
        rows += f'<tr style="background:{COLORS["light_bg"]}; font-weight:bold;">{sub_cells}</tr>'

    return section_header('시설별 객실 분석 (확정 기준)') + table_wrap('<thead>' + heads + '</thead><tbody>' + rows + '</tbody>')

def build_adr(df_conf, df_all):
    """ADR 분석"""
    total_sales = int(df_conf['판매가'].sum())
    total_nights = int(df_conf['박수'].sum())

    # ADR 요약 카드
    fac_adrs = df_conf.groupby('시설').apply(
        lambda g: adr(int(g['판매가'].sum()), int(g['박수'].sum()))
    )
    pkg_adrs = df_conf.groupby('패키지명_clean').apply(
        lambda g: adr(int(g['판매가'].sum()), int(g['박수'].sum()))
    )

    card_cells = (f'<td style="padding:10px 8px; text-align:center; border-right:1px solid {COLORS["border"]};">'
                  f'<div style="font-size:11px; color:#666; margin-bottom:4px;">전체 ADR</div>'
                  f'<div style="font-size:16px; font-weight:bold; color:{COLORS["dark"]};">'
                  f'{fmt(adr(total_sales,total_nights))}'
                  f'<span style="font-size:11px; color:#666; font-weight:normal;"> 원/박</span></div></td>')
    for name, v in list(fac_adrs.items())[:4]:
        card_cells += (f'<td style="padding:10px 8px; text-align:center; border-right:1px solid {COLORS["border"]};">'
                       f'<div style="font-size:11px; color:#666; margin-bottom:4px;">{name} ADR</div>'
                       f'<div style="font-size:16px; font-weight:bold; color:{COLORS["dark"]};">'
                       f'{fmt(v)}'
                       f'<span style="font-size:11px; color:#666; font-weight:normal;"> 원/박</span></div></td>')

    adr_card = (f'<table width="100%" cellpadding="0" cellspacing="0" border="0" '
                f'style="width:100%; border-collapse:collapse; border:1px solid {COLORS["border"]}; margin-top:4px;">'
                f'<tr>{card_cells}</tr></table>')

    # 패키지별 ADR 테이블
    pkg_heads = ('<thead><tr>'
                 + th('패키지', extra='width:120px;')
                 + th('건 수') + th('박 수') + th('확정 매출 (원)')
                 + th('ADR (원/박)') + th('건당 평균단가') + th('매출비중') + th('취소율')
                 + '</tr></thead>')

    pkg_rows = ''
    all_pkgs = df_all.groupby('패키지명_clean')
    for pkg_name, g in df_conf.groupby('패키지명_clean'):
        sales = int(g['판매가'].sum()); nights = int(g['박수'].sum()); cnt = len(g)
        try:
            gall = all_pkgs.get_group(pkg_name)
            x_cnt = len(gall[gall['상태'] == '취소'])
            cr = f"{round(x_cnt/(len(gall))*100,1)}%" if len(gall) > 0 else '-'
        except KeyError:
            x_cnt, cr = 0, '-'
        pct = f"{round(sales/total_sales*100,1)}%"
        cells = (td_l(pkg_name)
                 + td_r(fmt(cnt)) + td_r(fmt(nights)) + td_r(fmt(sales))
                 + td_r(fmt(adr(sales,nights)), bold=True)
                 + td_r(fmt(round(sales/cnt) if cnt else 0))
                 + td_c(pct) + td_c(cr, red=True))
        pkg_rows += f'<tr>{cells}</tr>'

    # 합계
    tx = len(df_all[df_all['상태']=='취소'])
    ttot = len(df_all)
    pkg_rows += (f'<tr style="background:{COLORS["light_bg"]}; font-weight:bold;">'
                 + td_l('합계')
                 + td_r(fmt(len(df_conf))) + td_r(fmt(total_nights)) + td_r(fmt(total_sales))
                 + td_r(fmt(adr(total_sales,total_nights)), bold=True)
                 + td_r(fmt(round(total_sales/len(df_conf)) if len(df_conf) else 0))
                 + td_c('100.0%')
                 + td_c(f"{round(tx/ttot*100,1)}%" if ttot else '-', red=True)
                 + '</tr>')

    pkg_table = (f'<p style="font-size:11px; font-weight:bold; color:#444; margin:10px 0 4px; padding-left:2px;">패키지별 ADR</p>'
                 + table_wrap('<thead>' + pkg_heads + '</thead><tbody>' + pkg_rows + '</tbody>'))

    # 객실별 ADR (ADR 높은 순)
    room_adr_df = df_conf.groupby('객실명_clean').agg(
        cnt=('판매가','count'), nights=('박수','sum'), sales=('판매가','sum')
    ).reset_index()
    room_adr_df['ADR'] = room_adr_df.apply(lambda r: adr(int(r['sales']), int(r['nights'])), axis=1)
    room_adr_df = room_adr_df.sort_values('ADR', ascending=False)

    room_heads = ('<thead><tr>'
                  + th('객실명', extra='width:180px;')
                  + th('건 수') + th('박 수')
                  + th('확정 매출 (원)') + th('ADR (원/박)') + th('건당 평균단가') + th('비중')
                  + '</tr></thead>')
    room_rows = ''
    for _, r in room_adr_df.iterrows():
        pct = f"{round(r['sales']/total_sales*100,1)}%"
        cells = (td_l(r['객실명_clean'], small=True)
                 + td_r(fmt(int(r['cnt']))) + td_r(fmt(int(r['nights'])))
                 + td_r(fmt(int(r['sales'])))
                 + td_r(fmt(int(r['ADR'])), bold=True)
                 + td_r(fmt(round(r['sales']/r['cnt']) if r['cnt'] else 0))
                 + td_c(pct))
        room_rows += f'<tr>{cells}</tr>'

    room_table = (f'<p style="font-size:11px; font-weight:bold; color:#444; margin:10px 0 4px; padding-left:2px;">객실별 ADR (ADR 높은 순)</p>'
                  + table_wrap('<thead>' + room_heads + '</thead><tbody>' + room_rows + '</tbody>'))

    return section_header('ADR 분석 (확정 기준)') + adr_card + pkg_table + room_table

def build_waitlist(df):
    """대기예약 vs 일반예약 비교 (대기예약 있을 때만)"""
    if not df['대기예약여부'].any():
        return ''

    wait = df[df['대기예약여부']]
    normal = df[~df['대기예약여부']]

    def rate(d):
        total = len(d)
        conf = len(d[d['상태'] == '확정'])
        return conf, total, f"{round(conf/total*100,1)}%" if total > 0 else '-'

    wc, wt, wr = rate(wait)
    nc, nt, nr = rate(normal)

    heads = ('<thead><tr>'
             + th('구분') + th('전체 건 수') + th('확정 건 수') + th('전환(확정)율')
             + '</tr></thead>')
    rows = (f'<tr>'
            + td_c('대기예약') + td_r(fmt(wt)) + td_r(fmt(wc)) + td_c(wr, red=True)
            + '</tr>'
            + f'<tr>'
            + td_c('일반예약') + td_r(fmt(nt)) + td_r(fmt(nc)) + td_c(nr)
            + '</tr>')

    note = (f'<p style="font-size:11px; color:#666; margin:5px 0 0; padding-left:2px;">'
            f'※ 대기예약 {fmt(wt)}건 중 {fmt(wc)}건 확정 전환(전환율 {wr}), '
            f'일반예약 대비 전환율 현저히 낮음</p>')

    return section_header('대기예약 vs 일반예약 전환 비교') + table_wrap('<thead>' + heads + '</thead><tbody>' + rows + '</tbody>') + note

# ────────────────────────────────────────────────
# 5. HTML 전체 조립
# ────────────────────────────────────────────────

def build_html(df, title, sale_period, stay_period):
    conf = df[df['상태'] == '확정']
    canc = df[df['상태'] == '취소']

    a_total = agg(df)
    cr = cancel_rate(a_total)

    # 시설별 집계
    facilities = sorted(df['시설'].unique())
    fac_dict = {}
    for fac in facilities:
        fac_dict[fac] = agg(df[df['시설'] == fac])

    # 섹션 조립
    sections = ''
    sections += build_kpi(a_total, cr)
    sections += '<div style="padding:0 16px 16px;">'
    sections += build_overview(a_total, fac_dict)
    sections += build_package(conf, df)
    sections += build_monthly(df)
    sections += build_daily(df)
    sections += build_weekday(conf)
    sections += build_waitlist(df)
    sections += build_room_analysis(df, conf, canc, facility_order=facilities)
    sections += build_adr(conf, df)
    sections += '</div>'

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<title>{title}</title>
</head>
<body style="margin:0; padding:0; font-family:Arial,sans-serif; font-size:13px; color:#222; background:#ffffff;">
<table width="600" cellpadding="0" cellspacing="0" border="0"
  style="width:600px; max-width:600px; margin:0; background:#ffffff; border-radius:8px; overflow:hidden;">
  <tr>
    <td style="background:{COLORS['dark']}; color:#ffffff; padding:10px 16px;
      font-size:14px; font-weight:bold; border-radius:6px 6px 0 0;">{title}</td>
  </tr>
  <tr>
    <td style="background:{COLORS['mid']}; color:#d6e8f7; padding:6px 16px; font-size:12px;">
      판매기간: {sale_period} &nbsp;|&nbsp; 투숙기간: {stay_period}
    </td>
  </tr>
  <tr><td>{sections}</td></tr>
</table>
</body>
</html>"""
    return html

# ────────────────────────────────────────────────
# 7. 엑셀 리포트 생성
# ────────────────────────────────────────────────

from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers as xl_numbers)
from openpyxl.utils import get_column_letter

# 색상 팔레트
XL = {
    'dark':      '1E4D7B',
    'mid':       '2E6DA4',
    'light_bg':  'EEF3F8',
    'col_bg':    'DCE8F4',
    'header_bg': 'F0F4F8',
    'red':       'C0392B',
    'white':     'FFFFFF',
    'border':    'B0BAC8',
    'gray_text': '666666',
}

def xlfill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def xlfont(bold=False, color='000000', size=10, name='Arial'):
    return Font(bold=bold, color=color, size=size, name=name)

def xlborder(all_sides=True):
    s = Side(style='thin', color=XL['border'])
    if all_sides:
        return Border(left=s, right=s, top=s, bottom=s)
    return Border(bottom=s)

def apply_outer_border(ws, min_row, min_col, max_row, max_col):
    """병합 셀 영역의 외곽 4면에 테두리를 완전히 적용"""
    s = Side(style='thin', color=XL['border'])
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row, col)
            left   = s if col == min_col else None
            right  = s if col == max_col else None
            top    = s if row == min_row else None
            bottom = s if row == max_row else None
            existing = cell.border
            cell.border = Border(
                left   = left   or existing.left,
                right  = right  or existing.right,
                top    = top    or existing.top,
                bottom = bottom or existing.bottom,
            )

def apply_full_border(ws, min_row, min_col, max_row, max_col):
    """모든 셀에 전체 테두리 적용 (그리드)"""
    s = Side(style='thin', color=XL['border'])
    b = Border(left=s, right=s, top=s, bottom=s)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).border = b


def xlalign(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def style_header(cell, dark=False, col=False):
    cell.fill = xlfill(XL['dark'] if dark else (XL['col_bg'] if col else XL['header_bg']))
    cell.font = xlfont(bold=True, color=XL['white'] if dark else XL['dark'], size=10)
    cell.border = xlborder()
    cell.alignment = xlalign()

def style_data(cell, bold=False, red=False, align='right'):
    cell.font = xlfont(bold=bold, color=XL['red'] if red else '000000')
    cell.border = xlborder()
    cell.alignment = xlalign(h=align)

def style_total(cell, align='right'):
    cell.fill = xlfill(XL['light_bg'])
    cell.font = xlfont(bold=True)
    cell.border = xlborder()
    cell.alignment = xlalign(h=align)

def style_section(cell, text):
    cell.value = text
    cell.fill = xlfill(XL['light_bg'])
    cell.font = xlfont(bold=True, color=XL['dark'], size=10)
    cell.alignment = xlalign(h='left')
    cell.border = Border(left=Side(style='thick', color=XL['mid']))

def num_fmt(cell, fmt='#,##0'):
    cell.number_format = fmt

def set_val(cell, val, bold=False, red=False, align='right', fmt='#,##0', center=False):
    cell.value = val
    cell.font = xlfont(bold=bold, color=XL['red'] if red else '000000')
    cell.border = xlborder()
    if center or isinstance(val, str):
        cell.alignment = xlalign(h='center')
    else:
        cell.alignment = xlalign(h=align)
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        cell.number_format = fmt

def merge_h(ws, row, col_start, col_end, val, dark=False, col_color=False, col=False):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=val)
    style_header(cell, dark=dark, col=col_color or col)
    return cell


# ── 시트 ① 종합 요약 ──────────────────────────────
def build_sheet1(ws, df, title, sale_period, stay_period):
    conf = df[df['상태'] == '확정']
    canc = df[df['상태'] == '취소']

    DOW_KR = {'Monday':'월','Tuesday':'화','Wednesday':'수','Thursday':'목',
               'Friday':'금','Saturday':'토','Sunday':'일'}
    DOW_ORDER = ['월','화','수','목','금','토','일']

    # 열 너비 A~N
    for ci, w in {1:2,2:14,3:8,4:8,5:16,6:8,7:8,8:16,9:8,10:8,11:16,12:8,13:8,14:16}.items():
        ws.column_dimensions[get_column_letter(ci)].width = w

    r = 1
    # ── 타이틀
    ws.merge_cells(f'B{r}:K{r}')
    c = ws.cell(r, 2, title)
    c.fill = xlfill(XL['dark']); c.font = xlfont(bold=True, color=XL['white'], size=13)
    c.alignment = xlalign(h='center'); ws.row_dimensions[r].height = 24
    apply_outer_border(ws, r, 2, r, 11)

    r += 1
    ws.merge_cells(f'B{r}:K{r}')
    c = ws.cell(r, 2, f'판매기간: {sale_period}   |   투숙기간: {stay_period}')
    c.fill = xlfill(XL['mid']); c.font = xlfont(color=XL['white'], size=10)
    c.alignment = xlalign(h='center')
    apply_outer_border(ws, r, 2, r, 11)

    r += 2
    # ── KPI (B~K, 각 2칸씩 5개)
    kpi_items = [
        ('확정 건수(확정+완료)', len(conf),                      '#,##0', False),
        ('확정 박수',            int(conf['박수'].sum()),         '#,##0', False),
        ('확정 매출액(원)',       int(conf['판매가'].sum()),       '#,##0', False),
        ('취소 건수',            int(canc.shape[0]),              '#,##0', False),
        ('취소율',               cancel_rate(agg(df)) / 100,     '0.0%',  True),
    ]
    for i, (lbl, val, fmt_str, red) in enumerate(kpi_items):
        col = 2 + i * 2
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
        lc = ws.cell(r, col, lbl)
        lc.fill = xlfill(XL['header_bg']); lc.font = xlfont(color=XL['gray_text'], size=9)
        lc.alignment = xlalign(h='center')
        apply_outer_border(ws, r, col, r, col+1)

        ws.merge_cells(start_row=r+1, start_column=col, end_row=r+1, end_column=col+1)
        vc = ws.cell(r+1, col, val)
        vc.font = xlfont(bold=True, color=XL['red'] if red else XL['dark'], size=14)
        vc.alignment = xlalign(h='center')
        vc.number_format = fmt_str
        apply_outer_border(ws, r+1, col, r+1, col+1)
        ws.row_dimensions[r+1].height = 26
    r += 3

    # ── 전체 예약 현황 (B~K)
    ws.merge_cells(f'B{r}:K{r}')
    c = ws.cell(r, 2, '▶  전체 예약 현황')
    c.fill = xlfill(XL['light_bg']); c.font = xlfont(bold=True, color=XL['dark'], size=10)
    c.alignment = xlalign(h='left')
    apply_outer_border(ws, r, 2, r, 11)
    c.border = Border(left=Side(style='thick', color=XL['mid']),
                      right=Side(style='thin', color=XL['border']),
                      top=Side(style='thin', color=XL['border']),
                      bottom=Side(style='thin', color=XL['border']))
    r += 1

    # 헤더 row1: 구분(B, rowspan2) + 전체(C:E) + 확정(F:H) + 취소(I:K)
    ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
    ws.cell(r, 2, '구분')
    apply_full_border(ws, r, 2, r+1, 2)
    ws.cell(r, 2).fill = xlfill(XL['header_bg'])
    ws.cell(r, 2).font = xlfont(bold=True, color=XL['dark'])
    ws.cell(r, 2).alignment = xlalign()

    for label, sc in [('전체 예약', 3), ('확정 예약', 6), ('취소 예약', 9)]:
        ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=sc+2)
        ws.cell(r, sc, label)
        ws.cell(r, sc).fill = xlfill(XL['col_bg'])
        ws.cell(r, sc).font = xlfont(bold=True, color=XL['dark'])
        ws.cell(r, sc).alignment = xlalign()
        apply_outer_border(ws, r, sc, r, sc+2)
    r += 1

    # 헤더 row2: 건수/박수/매출액 x3
    for sc in [3, 6, 9]:
        for offset, lbl in enumerate(['건 수', '박 수', '매출액 (원)']):
            c = ws.cell(r, sc+offset, lbl)
            c.fill = xlfill(XL['header_bg']); c.font = xlfont(bold=True, color=XL['dark'])
            c.alignment = xlalign(); c.border = xlborder()
    r += 1

    # 데이터 행
    a = agg(df)
    facilities = sorted(df['시설'].unique())
    for name, fa, is_total in [('전체', a, True)] + [(f, agg(df[df['시설']==f]), False) for f in facilities]:
        vals = [name, fa['전체건'], fa['전체박'], fa['전체매출'],
                fa['확정건'], fa['확정박'], fa['확정매출'],
                fa['취소건'], fa['취소박'], fa['취소매출']]
        for ci, v in enumerate(vals, 2):
            c = ws.cell(r, ci, v)
            if is_total: style_total(c, align='center' if ci==2 else 'right')
            else: style_data(c, align='center' if ci==2 else 'right')
            if isinstance(v, int): c.number_format = '#,##0'
        r += 1
    r += 1

    # ── 패키지별 현황 (B~J)
    ws.merge_cells(f'B{r}:J{r}')
    c = ws.cell(r, 2, '▶  패키지별 분석 (확정 기준)')
    c.fill = xlfill(XL['light_bg']); c.font = xlfont(bold=True, color=XL['dark'], size=10)
    c.alignment = xlalign(h='left')
    apply_outer_border(ws, r, 2, r, 10)
    c.border = Border(left=Side(style='thick', color=XL['mid']),
                      right=Side(style='thin', color=XL['border']),
                      top=Side(style='thin', color=XL['border']),
                      bottom=Side(style='thin', color=XL['border']))
    r += 1

    for ci, h in enumerate(['패키지','건 수','박 수','매출액 (원)','ADR(박)','ADR(건)','취소건','취소율','매출비중'], 2):
        c = ws.cell(r, ci, h)
        c.fill = xlfill(XL['header_bg']); c.font = xlfont(bold=True, color=XL['dark'])
        c.alignment = xlalign(); c.border = xlborder()
    r += 1

    total_sales = int(conf['판매가'].sum())
    all_pkgs = df.groupby('패키지명_clean')
    for pkg_name, g in conf.groupby('패키지명_clean'):
        sales = int(g['판매가'].sum()); nights = int(g['박수'].sum()); cnt = len(g)
        try:
            gall = all_pkgs.get_group(pkg_name)
            x_cnt = len(gall[gall['상태']=='취소'])
            cr_val = round(x_cnt/len(gall), 3)
        except: x_cnt=0; cr_val=0
        pct = round(sales/total_sales, 3) if total_sales else 0
        for ci, v in enumerate([pkg_name,cnt,nights,sales,adr(sales,nights),
                                  round(sales/cnt) if cnt else 0,x_cnt,cr_val,pct], 2):
            c = ws.cell(r, ci, v)
            style_data(c, align='left' if ci==2 else 'right')
            if ci in (3,4,5,6,7): c.number_format = '#,##0'
            if ci == 9: c.number_format = '0.0%'; c.font = xlfont(color=XL['red'])
            if ci == 10: c.number_format = '0.0%'
        r += 1

    ts=int(conf['판매가'].sum()); tn=int(conf['박수'].sum()); tc=len(conf)
    tx=len(df[df['상태']=='취소']); ttot=len(df)
    tcr=round(tx/ttot,3) if ttot else 0
    for ci, v in enumerate(['합계',tc,tn,ts,adr(ts,tn),round(ts/tc) if tc else 0,tx,tcr,1.0], 2):
        c = ws.cell(r, ci, v); style_total(c, align='left' if ci==2 else 'right')
        if ci in (3,4,5,6,7): c.number_format = '#,##0'
        if ci == 9: c.number_format = '0.0%'; c.font = xlfont(bold=True, color=XL['red'])
        if ci == 10: c.number_format = '0.0%'
    r += 2

    # ── 월별 체크인
    df_m = df.copy(); df_m['월'] = df_m['체크인'].dt.to_period('M')
    months = sorted(df_m['월'].unique())
    total_cols = 2 + len(months)*3 + 3
    max_col = total_cols

    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
    c = ws.cell(r, 2, '▶  체크인 월별 현황 (확정 기준)')
    c.fill = xlfill(XL['light_bg']); c.font = xlfont(bold=True, color=XL['dark'], size=10)
    c.alignment = xlalign(h='left')
    apply_outer_border(ws, r, 2, r, max_col)
    c.border = Border(left=Side(style='thick', color=XL['mid']),
                      right=Side(style='thin', color=XL['border']),
                      top=Side(style='thin', color=XL['border']),
                      bottom=Side(style='thin', color=XL['border']))
    r += 1

    ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
    ws.cell(r, 2, '구분')
    apply_full_border(ws, r, 2, r+1, 2)
    ws.cell(r, 2).fill = xlfill(XL['header_bg']); ws.cell(r, 2).font = xlfont(bold=True, color=XL['dark'])
    ws.cell(r, 2).alignment = xlalign()
    ws.cell(r+1, 2).fill = xlfill(XL['header_bg'])

    col = 3
    for m in months:
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+2)
        ws.cell(r, col, f'{m.year}년 {m.month:02d}월')
        ws.cell(r, col).fill = xlfill(XL['col_bg']); ws.cell(r, col).font = xlfont(bold=True, color=XL['dark'])
        ws.cell(r, col).alignment = xlalign()
        apply_outer_border(ws, r, col, r, col+2)
        col += 3
    ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+2)
    ws.cell(r, col, '합계')
    ws.cell(r, col).fill = xlfill(XL['col_bg']); ws.cell(r, col).font = xlfont(bold=True, color=XL['dark'])
    ws.cell(r, col).alignment = xlalign()
    apply_outer_border(ws, r, col, r, col+2)
    r += 1

    col = 3
    for _ in range(len(months)+1):
        for lbl in ['건','박','매출 (원)']:
            c = ws.cell(r, col, lbl)
            c.fill = xlfill(XL['header_bg']); c.font = xlfont(bold=True, color=XL['dark'])
            c.alignment = xlalign(); c.border = xlborder()
            col += 1
    r += 1

    for sub_df, name in [(df_m,'전체'),(df_m[df_m['상태']=='확정'],'확정'),(df_m[df_m['상태']=='취소'],'취소')]:
        c = ws.cell(r, 2, name); style_data(c, align='center')
        col = 3
        for m in months:
            g = sub_df[sub_df['월']==m]
            for v in [len(g), int(g['박수'].sum()), int(g['판매가'].sum())]:
                c = ws.cell(r, col, v); style_data(c); c.number_format='#,##0'; col+=1
        for v in [len(sub_df), int(sub_df['박수'].sum()), int(sub_df['판매가'].sum())]:
            c = ws.cell(r, col, v); style_total(c); c.number_format='#,##0'; col+=1
        r += 1
    r += 1

    # ── 일별 판매 현황 (B~G)
    ws.merge_cells(f'B{r}:G{r}')
    c = ws.cell(r, 2, f'▶  일별 판매 현황 (판매기간: {sale_period})')
    c.fill = xlfill(XL['light_bg']); c.font = xlfont(bold=True, color=XL['dark'], size=10)
    c.alignment = xlalign(h='left')
    apply_outer_border(ws, r, 2, r, 7)
    c.border = Border(left=Side(style='thick', color=XL['mid']),
                      right=Side(style='thin', color=XL['border']),
                      top=Side(style='thin', color=XL['border']),
                      bottom=Side(style='thin', color=XL['border']))
    r += 1

    for ci, h in enumerate(['날짜','요일','건 수','박 수','매출액 (원)','누적 매출액 (원)'], 2):
        c = ws.cell(r, ci, h)
        c.fill = xlfill(XL['header_bg']); c.font = xlfont(bold=True, color=XL['dark'])
        c.alignment = xlalign(); c.border = xlborder()
    r += 1

    df_d = df.copy(); df_d['예약일자'] = df_d['예약일'].dt.date
    daily = df_d.groupby('예약일자').agg(건수=('판매가','count'),박수=('박수','sum'),매출=('판매가','sum')).reset_index().sort_values('예약일자')
    cum = 0
    for _, row_d in daily.iterrows():
        cum += int(row_d['매출'])
        dow = DOW_KR.get(pd.Timestamp(row_d['예약일자']).day_name(), '')
        for ci, v in enumerate([row_d['예약일자'].strftime('%m-%d'),dow,
                                  int(row_d['건수']),int(row_d['박수']),int(row_d['매출']),cum], 2):
            c = ws.cell(r, ci, v)
            style_data(c, align='center' if ci<=3 else 'right')
            if ci >= 4: c.number_format = '#,##0'
        r += 1

    for ci, v in enumerate(['합계','',int(daily['건수'].sum()),int(daily['박수'].sum()),int(daily['매출'].sum()),''], 2):
        c = ws.cell(r, ci, v); style_total(c, align='center' if ci<=3 else 'right')
        if isinstance(v, int): c.number_format = '#,##0'
    r += 2

    # ── 요일별 현황 (B~G)
    ws.merge_cells(f'B{r}:G{r}')
    c = ws.cell(r, 2, '▶  체크인 요일별 현황 (확정 기준)')
    c.fill = xlfill(XL['light_bg']); c.font = xlfont(bold=True, color=XL['dark'], size=10)
    c.alignment = xlalign(h='left')
    apply_outer_border(ws, r, 2, r, 7)
    c.border = Border(left=Side(style='thick', color=XL['mid']),
                      right=Side(style='thin', color=XL['border']),
                      top=Side(style='thin', color=XL['border']),
                      bottom=Side(style='thin', color=XL['border']))
    r += 1

    for ci, h in enumerate(['요일','건 수','박 수','매출액 (원)','ADR (원/박)','비중 (매출)'], 2):
        c = ws.cell(r, ci, h)
        c.fill = xlfill(XL['header_bg']); c.font = xlfont(bold=True, color=XL['dark'])
        c.alignment = xlalign(); c.border = xlborder()
    r += 1

    conf_d = conf.copy()
    conf_d['요일_kr'] = conf_d['체크인'].dt.day_name().map(DOW_KR)
    total_s = int(conf['판매가'].sum()); total_n = int(conf['박수'].sum())
    dw = conf_d.groupby('요일_kr').agg(건수=('판매가','count'),박수=('박수','sum'),매출=('판매가','sum')).reindex(DOW_ORDER)
    for dow, row_w in dw.iterrows():
        if pd.isna(row_w['건수']): continue
        pct = round(row_w['매출']/total_s, 3) if total_s else 0
        for ci, v in enumerate([dow,int(row_w['건수']),int(row_w['박수']),
                                   int(row_w['매출']),adr(int(row_w['매출']),int(row_w['박수'])),pct], 2):
            c = ws.cell(r, ci, v); style_data(c, align='center' if ci==2 else 'right')
            if ci in (3,4,5): c.number_format = '#,##0'
            if ci == 7: c.number_format = '0.0%'
        r += 1

    for ci, v in enumerate(['합계',len(conf),total_n,total_s,adr(total_s,total_n),1.0], 2):
        c = ws.cell(r, ci, v); style_total(c, align='center' if ci==2 else 'right')
        if ci in (3,4,5): c.number_format = '#,##0'
        if ci == 7: c.number_format = '0.0%'


# ── 시트 ② 시설별 객실 분석 ────────────────────────
def build_sheet2(ws, df):
    conf = df[df['상태'] == '확정']
    canc = df[df['상태'] == '취소']

    col_widths = [2, 20, 8, 8, 14, 8, 8, 14, 8, 8, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    facilities = sorted(df['시설'].unique())
    fac_colors = [XL['dark'], XL['mid'], '3A7EBF']

    for fi, fac in enumerate(facilities):
        fa_all = df[df['시설'] == fac]
        fa_conf = conf[conf['시설'] == fac]
        fa_canc = canc[canc['시설'] == fac]

        # 시설 헤더
        ws.merge_cells(f'B{r}:K{r}')
        c = ws.cell(r, 2, f'▸  {fac}')
        c.fill = xlfill(fac_colors[fi % len(fac_colors)])
        c.font = xlfont(bold=True, color=XL['white'], size=11)
        c.alignment = xlalign(h='left')
        r += 1

        # 컬럼 헤더
        ws.cell(r, 2, '객실명'); style_header(ws.cell(r, 2))
        ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
        for label, sc in [('전체 예약', 3), ('확정 예약', 6), ('취소 예약', 9)]:
            merge_h(ws, r, sc, sc+2, label, col=True)
            ws.merge_cells(start_row=r, start_column=sc, end_row=r, end_column=sc+2)
        r += 1
        for col in range(3, 12):
            lbl = ['건','박','매출액 (원)'][(col-3) % 3]
            c = ws.cell(r, col, lbl); style_header(c)
        r += 1

        # 객실별
        for room in sorted(fa_all['객실명_clean'].unique()):
            ra = fa_all[fa_all['객실명_clean'] == room]
            rc = fa_conf[fa_conf['객실명_clean'] == room]
            rx = fa_canc[fa_canc['객실명_clean'] == room]
            vals = [room,
                    len(ra), int(ra['박수'].sum()), int(ra['판매가'].sum()),
                    len(rc), int(rc['박수'].sum()), int(rc['판매가'].sum()),
                    len(rx), int(rx['박수'].sum()), int(rx['판매가'].sum())]
            for ci, v in enumerate(vals, 2):
                c = ws.cell(r, ci, v)
                style_data(c, align='left' if ci==2 else 'right')
                if isinstance(v, int) and ci > 2: c.number_format = '#,##0'
            r += 1

        # 소계
        sub = [f'{fac} 소계',
               len(fa_all), int(fa_all['박수'].sum()), int(fa_all['판매가'].sum()),
               len(fa_conf), int(fa_conf['박수'].sum()), int(fa_conf['판매가'].sum()),
               len(fa_canc), int(fa_canc['박수'].sum()), int(fa_canc['판매가'].sum())]
        for ci, v in enumerate(sub, 2):
            c = ws.cell(r, ci, v); style_total(c, align='left' if ci==2 else 'right')
            if isinstance(v, int) and ci > 2: c.number_format = '#,##0'
        r += 2


# ── 시트 ③ 패키지×객실 매트릭스 ──────────────────────
def build_sheet3(ws, df):
    conf = df[df['상태'] == '확정']
    pkgs = sorted(conf['패키지명_clean'].unique())
    rooms = sorted(conf['객실명_clean'].unique())

    # 열 너비
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 22
    for i in range(3, 3 + len(pkgs) * 3 + 3):
        ws.column_dimensions[get_column_letter(i)].width = 9

    r = 1
    # 헤더
    ws.cell(r, 2, '객실명'); style_header(ws.cell(r, 2))
    ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=2)
    col = 3
    for pkg in pkgs:
        merge_h(ws, r, col, col+2, pkg, col=True)
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+2)
        ws.cell(r, col).alignment = xlalign(h='center', v='center', wrap=True)
        col += 3
    merge_h(ws, r, col, col+2, '소계', dark=True)
    ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+2)
    ws.row_dimensions[r].height = 48
    r += 1
    col = 3
    for _ in range(len(pkgs) + 1):
        for lbl in ['건','박','매출']:
            c = ws.cell(r, col, lbl); style_header(c); col += 1
    r += 1

    total_sales = int(conf['판매가'].sum())

    for room in rooms:
        ws.cell(r, 2, room).font = xlfont(); ws.cell(r, 2).border = xlborder(); ws.cell(r, 2).alignment = xlalign(h='left')
        col = 3
        row_total = [0, 0, 0]
        for pkg in pkgs:
            g = conf[(conf['객실명_clean'] == room) & (conf['패키지명_clean'] == pkg)]
            cnt = len(g); nts = int(g['박수'].sum()); sal = int(g['판매가'].sum())
            row_total[0] += cnt; row_total[1] += nts; row_total[2] += sal
            for v in [cnt, nts, sal]:
                c = ws.cell(r, col, v if v > 0 else ''); style_data(c)
                if isinstance(v, int) and v > 0: c.number_format = '#,##0'
                col += 1
        for v in row_total:
            c = ws.cell(r, col, v); style_total(c); c.number_format = '#,##0'; col += 1
        r += 1

    # 합계 행
    ws.cell(r, 2, '합계'); style_total(ws.cell(r, 2), align='center')
    col = 3
    for pkg in pkgs:
        g = conf[conf['패키지명_clean'] == pkg]
        for v in [len(g), int(g['박수'].sum()), int(g['판매가'].sum())]:
            c = ws.cell(r, col, v); style_total(c); c.number_format = '#,##0'; col += 1
    for v in [len(conf), int(conf['박수'].sum()), int(conf['판매가'].sum())]:
        c = ws.cell(r, col, v); style_total(c); c.number_format = '#,##0'; col += 1


# ── 시트 ④ ADR 분석 ──────────────────────────────
def build_sheet4(ws, df):
    conf = df[df['상태'] == '확정']
    total_sales = int(conf['판매가'].sum())
    total_nights = int(conf['박수'].sum())

    col_widths = [2, 20, 12, 10, 10, 16, 12, 12, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    # ADR 요약
    ws.merge_cells(f'B{r}:I{r}')
    c = ws.cell(r, 2, 'ADR (Average Daily Rate) 분석')
    c.fill = xlfill(XL['dark']); c.font = xlfont(bold=True, color=XL['white'], size=12)
    c.alignment = xlalign(h='center')
    r += 2

    # ADR 카드 (전체 + 시설별)
    fac_adrs = conf.groupby('시설').apply(lambda g: adr(int(g['판매가'].sum()), int(g['박수'].sum())))
    labels = ['전체 ADR'] + [f'{f} ADR' for f in fac_adrs.index]
    values = [adr(total_sales, total_nights)] + list(fac_adrs.values)
    for i, (lbl, val) in enumerate(zip(labels, values)):
        col = 2 + i * 2
        c = ws.cell(r, col, lbl)
        c.fill = xlfill(XL['header_bg']); c.font = xlfont(color=XL['gray_text'], size=9)
        c.alignment = xlalign(h='center'); c.border = xlborder()
        ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
        c2 = ws.cell(r+1, col, val)
        c2.font = xlfont(bold=True, color=XL['dark'], size=13)
        c2.alignment = xlalign(h='center'); c2.border = xlborder()
        c2.number_format = '#,##0"원/박"'
        ws.merge_cells(start_row=r+1, start_column=col, end_row=r+1, end_column=col+1)
    r += 3

    # 패키지별 ADR
    ws.merge_cells(f'B{r}:I{r}')
    style_section(ws.cell(r, 2), '▶  패키지별 ADR (확정 기준)')
    r += 1
    pkg_hdrs = ['패키지','건 수','박 수','확정 매출','ADR (원/박)','건당 평균단가','매출 비중','취소율']
    for ci, h in enumerate(pkg_hdrs, 2):
        c = ws.cell(r, ci, h); style_header(c)
    r += 1

    all_pkgs = df.groupby('패키지명_clean')
    for pkg_name, g in conf.groupby('패키지명_clean'):
        sales = int(g['판매가'].sum()); nights = int(g['박수'].sum()); cnt = len(g)
        try:
            gall = all_pkgs.get_group(pkg_name)
            x_cnt = len(gall[gall['상태']=='취소'])
            cr_val = round(x_cnt/len(gall), 3) if len(gall) else 0
        except: x_cnt=0; cr_val=0
        pct = round(sales/total_sales, 3) if total_sales else 0
        row_v = [pkg_name, cnt, nights, sales, adr(sales,nights), round(sales/cnt) if cnt else 0, pct, cr_val]
        for ci, v in enumerate(row_v, 2):
            c = ws.cell(r, ci, v); style_data(c, align='left' if ci==2 else 'right')
            if ci in (3,4,5,6): c.number_format = '#,##0'
            if ci == 7: c.number_format = '0.0%'
            if ci == 8: c.number_format = '0.0%'; c.font = xlfont(color=XL['red'])
        r += 1

    # 합계
    ts=int(conf['판매가'].sum()); tn=int(conf['박수'].sum()); tc=len(conf)
    tx=len(df[df['상태']=='취소']); ttot=len(df)
    row_t = ['합계', tc, tn, ts, adr(ts,tn), round(ts/tc) if tc else 0, 1.0, round(tx/ttot,3) if ttot else 0]
    for ci, v in enumerate(row_t, 2):
        c = ws.cell(r, ci, v); style_total(c, align='left' if ci==2 else 'right')
        if ci in (3,4,5,6): c.number_format = '#,##0'
        if ci == 7: c.number_format = '0.0%'
        if ci == 8: c.number_format = '0.0%'; c.font = xlfont(bold=True, color=XL['red'])
    r += 2

    # 객실별 ADR (높은 순)
    ws.merge_cells(f'B{r}:I{r}')
    style_section(ws.cell(r, 2), '▶  객실별 ADR (확정 기준 / ADR 높은 순)')
    r += 1
    room_hdrs = ['객실명','건 수','박 수','확정 매출','ADR (원/박)','건당 평균단가','비중']
    for ci, h in enumerate(room_hdrs, 2):
        c = ws.cell(r, ci, h); style_header(c)
    r += 1

    room_df = conf.groupby('객실명_clean').agg(
        cnt=('판매가','count'), nights=('박수','sum'), sales=('판매가','sum')
    ).reset_index()
    room_df['ADR_val'] = room_df.apply(lambda rw: adr(int(rw['sales']), int(rw['nights'])), axis=1)
    room_df = room_df.sort_values('ADR_val', ascending=False)

    for _, rw in room_df.iterrows():
        pct = round(rw['sales']/total_sales, 3) if total_sales else 0
        row_v = [rw['객실명_clean'], int(rw['cnt']), int(rw['nights']),
                 int(rw['sales']), int(rw['ADR_val']), round(rw['sales']/rw['cnt']) if rw['cnt'] else 0, pct]
        for ci, v in enumerate(row_v, 2):
            c = ws.cell(r, ci, v)
            style_data(c, bold=(ci==6), align='left' if ci==2 else 'right')
            if ci in (3,4,5,6,7): c.number_format = '#,##0'
            if ci == 8: c.number_format = '0.0%'
        r += 1


# ── 시트 ⑤ 체크인 일자별 ─────────────────────────
def build_sheet5(ws, df):
    conf = df[df['상태'] == '확정'].copy()
    conf['체크인일자'] = conf['체크인'].dt.date

    DOW_KR = {'Monday':'월','Tuesday':'화','Wednesday':'수','Thursday':'목',
               'Friday':'금','Saturday':'토','Sunday':'일'}

    col_widths = [2, 10, 4, 22, 14, 6, 6, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    ws.merge_cells(f'B{r}:H{r}')
    c = ws.cell(r, 2, '▶  체크인 일자별 현황 (확정 기준)')
    c.fill = xlfill(XL['light_bg']); c.font = xlfont(bold=True, color=XL['dark'], size=10)
    c.alignment = xlalign(h='left')
    apply_outer_border(ws, r, 2, r, 8)
    c.border = Border(left=Side(style='thick', color=XL['mid']),
                      right=Side(style='thin', color=XL['border']),
                      top=Side(style='thin', color=XL['border']),
                      bottom=Side(style='thin', color=XL['border']))
    r += 1

    for ci, h in enumerate(['체크인','요일','객실명','패키지','건 수','박 수','매출액 (원)'], 2):
        c = ws.cell(r, ci, h); style_header(c)
    r += 1

    grouped = conf.groupby(['체크인일자','객실명_clean','패키지명_clean']).agg(
        건수=('판매가','count'), 박수=('박수','sum'), 매출=('판매가','sum')
    ).reset_index().sort_values('체크인일자')

    prev_date = None
    for _, rw in grouped.iterrows():
        date_str = rw['체크인일자'].strftime('%m-%d') if rw['체크인일자'] != prev_date else ''
        dow = DOW_KR.get(pd.Timestamp(rw['체크인일자']).day_name(), '') if rw['체크인일자'] != prev_date else ''
        prev_date = rw['체크인일자']
        row_v = [date_str, dow, rw['객실명_clean'], rw['패키지명_clean'],
                 int(rw['건수']), int(rw['박수']), int(rw['매출'])]
        for ci, v in enumerate(row_v, 2):
            c = ws.cell(r, ci, v); style_data(c, align='center' if ci<=3 else ('left' if ci<=5 else 'right'))
            if ci >= 7: c.number_format = '#,##0'
        r += 1

    # 합계
    for ci, v in enumerate(['합계','','','', len(conf), int(conf['박수'].sum()), int(conf['판매가'].sum())], 2):
        c = ws.cell(r, ci, v); style_total(c, align='center' if ci<=5 else 'right')
        if isinstance(v, int): c.number_format = '#,##0'


# ── 엑셀 전체 생성 ─────────────────────────────
def build_excel(df, title, sale_period, stay_period, out_path):
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거

    sheet_defs = [
        ('① 종합 요약',       lambda ws: build_sheet1(ws, df, title, sale_period, stay_period)),
        ('② 시설별 객실 분석', lambda ws: build_sheet2(ws, df)),
        ('③ 패키지×객실 매트릭스', lambda ws: build_sheet3(ws, df)),
        ('④ ADR 분석',        lambda ws: build_sheet4(ws, df)),
        ('⑤ 체크인 일자별',   lambda ws: build_sheet5(ws, df)),
    ]

    for sheet_name, builder in sheet_defs:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.row_dimensions[1].height = 22
        builder(ws)

    wb.save(out_path)

